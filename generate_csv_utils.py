import os
import csv
from dotenv import load_dotenv

# Tentar importar openpyxl para suporte a XLS
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    Font = None
    Alignment = None
    PatternFill = None
    Border = None
    Side = None

# Carregar variáveis de ambiente
load_dotenv()
from datetime import datetime


def is_valid_title_record(linha):
    """
    Verifica se uma linha é um registro válido de título/boleto
    
    Args:
        linha (str): Linha do arquivo CNAB
        
    Returns:
        bool: True se for um registro válido de título
    """
    try:
        # Verificar tamanho mínimo
        if len(linha) < 267:
            return False
        
        # Verificar se tem número de documento válido (não vazio e não só zeros)
        n_documento = linha[116:131].strip()
        if not n_documento or n_documento == '000000000000000' or len(n_documento) < 3:
            return False
        
        # Verificar se tem valor válido (maior que zero)
        valor_str = linha[251:267].strip()
        if not valor_str or not valor_str.isdigit() or int(valor_str) <= 0:
            return False
        
        # Verificar se tem data válida (formato DDMMAA)
        data_str = linha[110:116].strip()
        if len(data_str) != 6 or not data_str.isdigit():
            return False
        
        # Validar se a data faz sentido
        dia = int(data_str[:2])
        mes = int(data_str[2:4])
        if dia < 1 or dia > 31 or mes < 1 or mes > 12:
            return False
        
        # Verificar se não é linha de header/trailer (geralmente começam com códigos específicos)
        # Headers/trailers geralmente têm padrões diferentes nos primeiros caracteres
        primeiro_char = linha[0] if linha else ''
        if primeiro_char in ['0', '9']:  # 0=header, 9=trailer em muitos formatos CNAB
            return False
        
        return True
        
    except Exception:
        return False


def extract_document_data(linha):
    """
    Extrai dados do documento de uma linha CNAB para geração de CSV
    
    Args:
        linha (str): Linha do arquivo CNAB
        
    Returns:
        dict: Dicionário com os dados extraídos (n_documento, valor, data_pagamento)
              ou None se não conseguir extrair os dados
    """
    try:
        # Primeiro verificar se é um registro válido de título
        if not is_valid_title_record(linha):
            return None
        
        # Extrair número do documento (posições 117-131)
        n_documento = linha[116:131].strip()
        
        # Extrair valor (posições 252-267)
        valor_str = linha[251:267].strip()
        if valor_str and valor_str.isdigit():
            # Dividir por 1000 para obter o formato correto (centavos)
            valor_int = int(valor_str)
            valor_formatado = f"{valor_int / 1000:.2f}".replace('.', ',')
        else:
            valor_formatado = "0,00"
        
        # Extrair data de vencimento (posições 111-116) formato DDMMAA
        data_str = linha[110:116].strip()
        if len(data_str) == 6:
            dia = data_str[:2]
            mes = data_str[2:4]
            ano = data_str[4:6]
            # Assumir que anos 00-30 são 2000-2030, e 31-99 são 1931-1999
            if int(ano) <= 30:
                ano_completo = f"20{ano}"
            else:
                ano_completo = f"19{ano}"
            data_formatada = f"{dia}/{mes}/{ano_completo}"
        else:
            data_formatada = ""
        
        return {
            'n_documento': n_documento,
            'valor': valor_formatado,
            'data_pagamento': data_formatada
        }
        
    except Exception as e:
        print(f"Erro ao extrair dados da linha: {str(e)}")
        return None


def generate_csv_from_cnab_lines(linhas_antecipadas, output_path):
    """
    Gera arquivo CSV a partir das linhas de operações antecipadas
    
    Args:
        linhas_antecipadas (list): Lista de linhas do arquivo CNAB com operações antecipadas
        output_path (str): Caminho onde salvar o arquivo CSV
        
    Returns:
        tuple: (bool, str) - (sucesso, mensagem)
    """
    try:
        dados_documentos = []
        
        # Processar cada linha
        for linha in linhas_antecipadas:
            dados = extract_document_data(linha)
            if dados and dados['n_documento']:  # Só adicionar se tiver documento
                dados_documentos.append(dados)
        
        if not dados_documentos:
            return False, "Nenhum documento válido encontrado"
        
        # Gerar CSV
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['n_documento', 'valor', 'data_pagamento']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            # Escrever cabeçalho
            writer.writeheader()
            
            # Escrever dados
            for dados in dados_documentos:
                writer.writerow(dados)
        
        return True, f"CSV gerado com sucesso: {len(dados_documentos)} registros salvos em {output_path}"
        
    except Exception as e:
        return False, f"Erro ao gerar CSV: {str(e)}"


def generate_xls_from_cnab_lines(linhas, xls_path):
    """
    Gera arquivo XLS a partir de linhas CNAB
    
    Args:
        linhas (list): Lista de linhas do arquivo CNAB
        xls_path (str): Caminho onde salvar o arquivo XLS
        
    Returns:
        tuple: (bool, str) - (sucesso, mensagem)
    """
    if not OPENPYXL_AVAILABLE:
        return False, "Biblioteca openpyxl não está instalada. Use 'pip install openpyxl' para instalar."
    
    try:
        dados_documentos = []
        
        # Processar cada linha (a validação é feita na função extract_document_data)
        for linha in linhas:
            dados = extract_document_data(linha)
            if dados and dados['n_documento']:  # Só adicionar se tiver documento
                dados_documentos.append(dados)
        
        if not dados_documentos:
            return False, "Nenhum documento válido encontrado"
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Boletos Antecipados"
        
        # Definir cabeçalhos
        headers = ['Número do Documento', 'Valor', 'Data de Pagamento']
        ws.append(headers)
        
        # Aplicar formatação ao cabeçalho
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Adicionar dados
        for i, dados in enumerate(dados_documentos, start=2):  # Começar na linha 2 (após cabeçalho)
            # Converter valor brasileiro (com vírgula) para float
            valor_str = dados['valor'].replace(',', '.')
            try:
                valor_float = float(valor_str)
            except ValueError:
                valor_float = 0.0
            
            ws.append([dados['n_documento'], valor_float, dados['data_pagamento']])
            
            # Formatar coluna de valor como moeda brasileira
            valor_cell = ws[f'B{i}']
            valor_cell.number_format = 'R$ #,##0.00'
            valor_cell.alignment = Alignment(horizontal='right')
            
            # Centralizar data
            data_cell = ws[f'C{i}']
            data_cell.alignment = Alignment(horizontal='center')
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 25  # Número do documento
        ws.column_dimensions['B'].width = 18  # Valor (mais espaço para moeda)
        ws.column_dimensions['C'].width = 18  # Data de pagamento
        
        # Adicionar bordas para melhor visualização
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar bordas em todas as células com dados
        for row in ws.iter_rows(min_row=1, max_row=len(dados_documentos)+1, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border
        
        # Salvar arquivo
        wb.save(xls_path)
        
        return True, f"XLS gerado com sucesso: {len(dados_documentos)} registros salvos em {xls_path}"
        
    except Exception as e:
        return False, f"Erro ao gerar XLS: {str(e)}"


def generate_output_for_antecipated_operations(arquivo_antecipado):
    """
    Gera arquivo de saída (CSV ou XLS) para operações antecipadas a partir de um arquivo .ret
    
    Args:
        arquivo_antecipado (str): Caminho para o arquivo .ret com operações antecipadas
        
    Returns:
        tuple: (bool, str, str) - (sucesso, mensagem, caminho_arquivo)
    """
    try:
        # Verificar se o arquivo existe
        if not os.path.exists(arquivo_antecipado):
            return False, f"Arquivo não encontrado: {arquivo_antecipado}", None
        
        # Obter formato de saída do .env (padrão: csv)
        output_format = os.getenv('OUTPUT_FORMAT', 'csv').lower()
        
        # Definir caminho do arquivo de saída
        nome_base = os.path.splitext(arquivo_antecipado)[0]
        if output_format == 'xls':
            output_path = f"{nome_base}.xlsx"
        else:
            output_path = f"{nome_base}.csv"
        
        # Ler arquivo
        try:
            with open(arquivo_antecipado, 'r', encoding='utf-8') as f:
                linhas = f.readlines()
        except UnicodeDecodeError:
            with open(arquivo_antecipado, 'r', encoding='latin-1') as f:
                linhas = f.readlines()
        
        # Limpar quebras de linha
        linhas = [linha.rstrip('\n') for linha in linhas]
        
        # Gerar arquivo baseado no formato
        if output_format == 'xls':
            sucesso, mensagem = generate_xls_from_cnab_lines(linhas, output_path)
        else:
            sucesso, mensagem = generate_csv_from_cnab_lines(linhas, output_path)
        
        if sucesso:
            return True, mensagem, output_path
        else:
            return False, mensagem, None
            
    except Exception as e:
        return False, f"Erro ao processar arquivo antecipado: {str(e)}", None


# Manter função original para compatibilidade
def generate_csv_for_antecipated_operations(arquivo_antecipado):
    """
    Função mantida para compatibilidade - usa a nova função com formato configurável
    """
    return generate_output_for_antecipated_operations(arquivo_antecipado)
